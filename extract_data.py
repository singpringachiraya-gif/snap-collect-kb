# -*- coding: utf-8 -*-
"""
Extracts the Snap&Collect master data from the source xlsx into clean JSON
for the searchable knowledge-base site (index.html).

    pip install openpyxl
    python extract_data.py
"""
import json, os, sys
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

SRC = os.environ.get("SNAP_EXCEL_SRC") or r"รวมข้อมูลการสะสมใบเสร็จ Snap&Collect (1).xlsx"
OUT = r"snap-collect-data.json"

wb = openpyxl.load_workbook(SRC, data_only=True)

def find_sheet(keyword, required=True):
    """Find first sheet whose name contains keyword (case-insensitive)."""
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return wb[name]
    if required:
        raise KeyError(f"No sheet containing '{keyword}'. Sheets: {wb.sheetnames}")
    return None

# --- Sheet: stores (contains "Y2026") ----------------------------------------
ws = find_sheet("Y2026")
stores = []
for row in ws.iter_rows(min_row=2, values_only=True):
    no, name, receipt_format, notes, category, extra, *_ = list(row) + [None] * 7
    if not name:
        continue
    stores.append({
        "no": no,
        "name": str(name).strip(),
        "receipt_format": str(receipt_format).strip() if receipt_format else "",
        "notes": str(notes).strip() if notes else "",
        "category": str(category).strip() if category else "",
        "extra": str(extra).strip() if extra else "",
    })

# --- Sheet: name mapping (optional — not present in all versions) ------------
ws2 = find_sheet("nam", required=False)
name_map = []
if ws2:
    for row in ws2.iter_rows(min_row=2, values_only=True):
        header_name, portal_name, note, *_ = list(row) + [None] * 4
        if not header_name and not portal_name:
            continue
        name_map.append({
            "header_name": str(header_name).strip() if header_name else "",
            "portal_name": str(portal_name).strip() if portal_name else "",
            "note": str(note).strip() if note else "",
        })

# --- Sheet: Reject cases (contains "Reject") ---------------------------------
ws3 = find_sheet("Reject", required=False)
rejects = []
if ws3:
    for row in ws3.iter_rows(min_row=2, values_only=True):
        case, reason, message, store_note, *_ = list(row) + [None] * 5
        if not case and not message:
            continue
        rejects.append({
            "case":       str(case).strip()       if case       else "",
            "reason":     str(reason).strip()     if reason     else "",
            "message":    str(message).strip()    if message    else "",
            "store_note": str(store_note).strip() if store_note else "",
        })

categories = sorted({s["category"] for s in stores if s["category"]})

with open(OUT, "w", encoding="utf-8") as f:
    json.dump({
        "stores": stores,
        "name_map": name_map,
        "rejects": rejects,
        "categories": categories,
    }, f, ensure_ascii=False, indent=1)

print(f"Wrote {len(stores)} stores, {len(name_map)} name-mapping rows, "
      f"{len(rejects)} reject cases, {len(categories)} categories -> {OUT}")
